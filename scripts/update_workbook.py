#!/usr/bin/env python3

"""
Update the investigation workbook and publish a GitHub-viewable CSV copy.

Reads:
    data/current_case.json

Writes:
    workbooks/Exposure-Tracking-Matrix.xlsx
    workbooks/Exposure-Tracking-Matrix.csv
"""

import csv
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

CASE_PATH = Path("data/current_case.json")
WORKBOOK_PATH = Path(
    "workbooks/Exposure-Tracking-Matrix.xlsx"
)
CSV_PATH = Path(
    "workbooks/Exposure-Tracking-Matrix.csv"
)

SHEET_NAME = "Investigations"

HEADERS = [
    "Date",
    "Case ID",
    "Campaign",
    "Classification",
    "Severity",
    "Priority",
    "Risk Score",
    "Confidence",
    "Evidence",
    "IOCs",
    "Assets",
    "Platform",
    "Vendor",
    "Network Zone",
    "Lead Analyst",
    "Status",
]

SEVERITY_WEIGHT = {
    "LOW": 1,
    "MODERATE": 2,
    "HIGH": 3,
    "CRITICAL": 4,
}

LEGACY_CAMPAIGN_NAMES = {
    "Operation Ashcroft": "Biomedical Evidence Review",
    "Operation Eclipse": (
        "Protected Research Access Investigation"
    ),
    "Operation Night Watch": "After-Hours Access Review",
    "Operation Chimera": "Specimen Integrity Investigation",
    "Operation Lazarus": (
        "Research System Recovery Assessment"
    ),
    "Operation Dead Signal": (
        "Medical Device Communications Review"
    ),
    "Operation Cold Harbor": (
        "Research Facility Security Assessment"
    ),
    "Operation Nightfall": (
        "Laboratory Network Exposure Review"
    ),
    "Operation Outbreak": "Biomedical Containment Incident",
    "Operation Black Eclipse": (
        "Coordinated Biomedical Systems Intrusion"
    ),
}

COLUMN_WIDTHS = {
    "A": 14,
    "B": 18,
    "C": 24,
    "D": 38,
    "E": 12,
    "F": 12,
    "G": 12,
    "H": 12,
    "I": 12,
    "J": 12,
    "K": 12,
    "L": 28,
    "M": 18,
    "N": 24,
    "O": 24,
    "P": 20,
}


def load_case() -> dict:
    """Load the active investigation record."""

    with CASE_PATH.open("r", encoding="utf-8") as file:
        data = json.load(file)

    if not isinstance(data, dict):
        raise ValueError(
            "data/current_case.json must contain a JSON object."
        )

    return data


def safe_int(value: object, default: int = 0) -> int:
    """Convert a value to an integer safely."""

    try:
        return int(float(str(value).strip()))

    except (TypeError, ValueError):
        return default


def normalized_risk_score(
    severity: object,
    confidence: object,
) -> int:
    """
    Calculate the standardized 0-100 workbook risk score.

    Severity controls the maximum score band:
        LOW       -> 0-25
        MODERATE  -> 0-50
        HIGH      -> 0-75
        CRITICAL  -> 0-100
    """

    severity_name = str(
        severity or "LOW"
    ).strip().upper()

    confidence_value = max(
        0,
        min(
            100,
            safe_int(confidence),
        ),
    )

    severity_factor = (
        SEVERITY_WEIGHT.get(
            severity_name,
            1,
        )
        / 4
    )

    return round(
        severity_factor
        * confidence_value
    )


def normalize_legacy_risk_scores(
    worksheet,
) -> int:
    """
    Convert legacy severity-weighted scores to the 0-100 scale.

    The old workbook formula was:
        severity weight * confidence

    A row is changed only when its current value exactly matches that
    legacy formula. Newer case-provided scores are therefore preserved.
    """

    header_map = {
        str(
            worksheet.cell(
                1,
                column_number,
            ).value
            or ""
        ).strip(): column_number
        for column_number in range(
            1,
            worksheet.max_column + 1,
        )
    }

    required_headers = {
        "Severity",
        "Risk Score",
        "Confidence",
    }

    if not required_headers.issubset(header_map):
        return 0

    severity_column = header_map["Severity"]
    risk_column = header_map["Risk Score"]
    confidence_column = header_map["Confidence"]

    changed_rows = 0

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        severity = str(
            worksheet.cell(
                row_number,
                severity_column,
            ).value
            or ""
        ).strip().upper()

        confidence = max(
            0,
            min(
                100,
                safe_int(
                    worksheet.cell(
                        row_number,
                        confidence_column,
                    ).value
                ),
            ),
        )

        current_score = safe_int(
            worksheet.cell(
                row_number,
                risk_column,
            ).value,
            default=-1,
        )

        severity_weight = SEVERITY_WEIGHT.get(
            severity
        )

        if severity_weight is None:
            continue

        legacy_score = (
            severity_weight
            * confidence
        )

        standardized_score = normalized_risk_score(
            severity,
            confidence,
        )

        if (
            current_score == legacy_score
            and current_score != standardized_score
        ):
            worksheet.cell(
                row_number,
                risk_column,
            ).value = standardized_score

            changed_rows += 1

    return changed_rows


def migrate_campaign_names(
    worksheet,
) -> tuple[int, bool]:
    """
    Rename the legacy Operation column and update historical campaign
    values to restrained investigative titles.

    Returns:
        A tuple containing:
            - the number of historical campaign values changed
            - whether the column header was renamed
    """

    header_map = {
        str(
            worksheet.cell(
                1,
                column_number,
            ).value
            or ""
        ).strip(): column_number
        for column_number in range(
            1,
            worksheet.max_column + 1,
        )
    }

    campaign_column = header_map.get(
        "Campaign"
    )

    header_renamed = False

    if campaign_column is None:
        campaign_column = header_map.get(
            "Operation"
        )

        if campaign_column is not None:
            worksheet.cell(
                1,
                campaign_column,
            ).value = "Campaign"

            header_renamed = True

    if campaign_column is None:
        return 0, header_renamed

    changed_rows = 0

    for row_number in range(
        2,
        worksheet.max_row + 1,
    ):
        cell = worksheet.cell(
            row_number,
            campaign_column,
        )

        current_name = str(
            cell.value
            or ""
        ).strip()

        replacement = LEGACY_CAMPAIGN_NAMES.get(
            current_name
        )

        if replacement and replacement != current_name:
            cell.value = replacement
            changed_rows += 1

    return changed_rows, header_renamed


def calculate_risk_score(case: dict) -> int:
    """
    Use the investigation's own 0-100 risk score when available.

    When no case score exists, calculate the same standardized 0-100
    workbook score used by the legacy-data normalization.
    """

    if case.get("risk_score") not in (None, ""):
        return max(
            0,
            min(
                100,
                safe_int(case.get("risk_score")),
            ),
        )

    return normalized_risk_score(
        case.get("severity"),
        case.get("confidence"),
    )


def open_workbook():
    """Open the existing workbook or create a new one."""

    WORKBOOK_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    if WORKBOOK_PATH.exists():
        workbook = load_workbook(WORKBOOK_PATH)
    else:
        workbook = Workbook()

    if SHEET_NAME in workbook.sheetnames:
        worksheet = workbook[SHEET_NAME]
    else:
        worksheet = workbook.active
        worksheet.title = SHEET_NAME

    if worksheet.max_row == 1:
        existing_header = [
            worksheet.cell(1, column).value
            for column in range(1, len(HEADERS) + 1)
        ]

        if not any(existing_header):
            worksheet.append(HEADERS)

    return workbook, worksheet


def build_row(case: dict) -> list[object]:
    """Build one workbook row from the current case."""

    return [
        case.get("date", ""),
        case.get("case_id", ""),
        case.get("operation", ""),
        case.get("classification", ""),
        case.get("severity", ""),
        case.get("priority", ""),
        calculate_risk_score(case),
        safe_int(case.get("confidence")),
        safe_int(case.get("evidence_count")),
        safe_int(case.get("ioc_count")),
        safe_int(case.get("affected_assets")),
        case.get("affected_platform", ""),
        case.get("vendor", ""),
        case.get("network_zone", ""),
        case.get("lead_analyst", ""),
        case.get("status", ""),
    ]


def upsert_case(
    worksheet,
    row_values: list[object],
) -> str:
    """
    Add the current case or update its existing row.

    This prevents duplicate workbook entries when the same workflow is
    rerun for an unchanged case.
    """

    case_id = str(row_values[1]).strip()
    target_row = None

    if case_id:
        for row_number in range(
            2,
            worksheet.max_row + 1,
        ):
            existing_case_id = str(
                worksheet.cell(
                    row_number,
                    2,
                ).value
                or ""
            ).strip()

            if existing_case_id == case_id:
                target_row = row_number
                break

    if target_row is None:
        worksheet.append(row_values)
        return "added"

    for column_number, value in enumerate(
        row_values,
        start=1,
    ):
        worksheet.cell(
            target_row,
            column_number,
        ).value = value

    return "updated"


def format_worksheet(worksheet) -> None:
    """Apply readable workbook formatting."""

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 24

    for column, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width


def export_csv(worksheet) -> None:
    """Write a GitHub-renderable CSV copy of the workbook sheet."""

    CSV_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with CSV_PATH.open(
        "w",
        newline="",
        encoding="utf-8",
    ) as file:
        writer = csv.writer(file)

        for row in worksheet.iter_rows(
            values_only=True,
        ):
            writer.writerow(
                [
                    "" if value is None else value
                    for value in row
                ]
            )


def main() -> None:
    case = load_case()
    workbook, worksheet = open_workbook()

    migrated_campaigns, header_renamed = (
        migrate_campaign_names(
            worksheet
        )
    )

    normalized_rows = normalize_legacy_risk_scores(
        worksheet
    )

    action = upsert_case(
        worksheet,
        build_row(case),
    )

    format_worksheet(worksheet)

    workbook.save(WORKBOOK_PATH)
    export_csv(worksheet)

    print(
        f"Workbook {action}: "
        f"{WORKBOOK_PATH}"
    )
    print(
        "Campaign header renamed: "
        f"{header_renamed}"
    )
    print(
        "Historical campaign names migrated: "
        f"{migrated_campaigns}"
    )
    print(
        "Legacy risk scores normalized: "
        f"{normalized_rows}"
    )
    print(
        f"GitHub CSV preview updated: "
        f"{CSV_PATH}"
    )


if __name__ == "__main__":
    main()
